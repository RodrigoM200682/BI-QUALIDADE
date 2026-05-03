# app.py — INDICADORES QUALIDADE RS (WEB) — visão simplificada
# Upgrade aplicado:
# - Sem senha de abertura
# - Mantém sempre o último upload salvo localmente
# - Filtro inicial por Ano e Mês
# - Modo de comparação entre dois períodos
# - Comparação permite selecionar um ou mais meses no Período 1
# - No comparativo, o Período 2 replica automaticamente os mesmos meses do Período 1
# - Gráfico 1: número de ocorrências
# - Gráfico 2: motivos das ocorrências
# - Em comparação, ambos os gráficos usam barras com cores distintas por período/ano
# - Exportação Excel inclui tabela, gráficos e dados de variação entre períodos
# - Excel: eixo X dos gráficos em 45° e aba RECORTE compacta
# - Abertura sempre em consulta normal, com todos os anos e meses da última base
# - Filtro de mês ajustado para seleção direta, sem opção (Todos), reduzindo necessidade de duplo clique
# - Versão dinâmica no formato RM_dd_mm_aaaa_HR, em itálico e fonte 10
# - Botão de exportação descrito como Gerar relatório
# - RECORTE com fonte/dimensões fixas e colunas auxiliares ocultas
# - Gráficos nativos das abas DADOS com eixo X correto e rótulos sobre barras
# - Recuperação automática reforçada da última base após inatividade
# - Campo Manual integrado ao aplicativo, com download do PDF de funcionamento

import io
import json
from datetime import datetime
from pathlib import Path

import pandas as pd
import streamlit as st
import plotly.express as px

from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import BarChart
from openpyxl.chart.reference import Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.drawing.image import Image as XLImage

# =========================================================
# Configuração geral
# =========================================================
APP_VERSION = datetime.now().strftime("RM_%d_%m_%Y_%HHR")
APP_NAME = "INDICADORES QUALIDADE RS"
DEFAULT_SHEET = "Sheet1"

LAST_DIR = Path(".last_input")
LAST_DIR.mkdir(exist_ok=True)
LAST_FILE = LAST_DIR / "last_excel.bin"
LAST_META = LAST_DIR / "last_excel_meta.json"
APP_STATE_FILE = LAST_DIR / "app_state_simplificado.json"

COL_CODIGO = "Código"
COL_TITULO = "Título"
COL_STATUS = "Status"
COL_DATA = "Data de emissão"
COL_MOTIVO = "Motivo Reclamação"
COL_RESP_OCORRENCIA = "Responsável"
COL_RESP_ANALISE = "Responsável da análise de causa"
COL_CATEGORIA = "Categoria"
COL_CLIENTE = "Cliente"
COL_SITUACAO = "Situação"

DATE_FMT_BR = "%d/%m/%Y"
MESES_ABREV = {
    1: "Jan", 2: "Fev", 3: "Mar", 4: "Abr", 5: "Mai", 6: "Jun",
    7: "Jul", 8: "Ago", 9: "Set", 10: "Out", 11: "Nov", 12: "Dez",
}
INV_MESES_ABREV = {v: k for k, v in MESES_ABREV.items()}

CORES_COMPARACAO = ["#1F5C4B", "#8BBF24"]
COR_NORMAL = "#1F5C4B"
COR_ATRASO = "#B42318"
COR_FUNDO = "#F4F7F2"
COR_FUNDO_2 = "#EAF1EA"
COR_TEXTO = "#173D35"
COR_CARD = "#FFFFFF"
COR_BORDA = "#D7E1D7"
COR_SIDEBAR = "#EEF4EA"
LOGO_FILE = Path(__file__).with_name("logo_brasilata.png")
DEFAULT_DATA_FILE = Path(__file__).with_name("Consultas_RNC_APP.xlsx")
MANUAL_FILE = Path(__file__).with_name("Manual_APP_Indicadores_Qualidade_V_RM_2026.pdf")


# =========================================================
# Persistência de upload e estado
# =========================================================
def _save_last_upload(xls_bytes: bytes, filename: str, sheet_name: str) -> None:
    """Salva a última base enviada em local dedicado e mantém uma cópia padrão no diretório do app.

    A cópia padrão funciona como redundância para quando a sessão do navegador expira
    ou quando o app reinicia e precisa localizar automaticamente a última base disponível.
    """
    try:
        LAST_DIR.mkdir(exist_ok=True)
        LAST_FILE.write_bytes(xls_bytes)
        LAST_META.write_text(
            json.dumps({"filename": filename or "ultimo.xlsx", "sheet": sheet_name or DEFAULT_SHEET}, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
    except Exception:
        pass

    # Redundância: mantém uma base padrão junto ao app para recuperação após inatividade/reinício.
    try:
        DEFAULT_DATA_FILE.write_bytes(xls_bytes)
    except Exception:
        pass


def _load_last_upload():
    """Carrega a última base conhecida sem depender de st.session_state.

    Ordem de busca:
    1) .last_input/last_excel.bin, salvo no último upload;
    2) Consultas_RNC_APP.xlsx no diretório do app, usado como fallback persistente do projeto.
    """
    try:
        if LAST_FILE.exists() and LAST_FILE.stat().st_size > 0:
            data = LAST_FILE.read_bytes()
            meta = {}
            if LAST_META.exists():
                meta = json.loads(LAST_META.read_text(encoding="utf-8"))
            return data, meta
    except Exception:
        pass

    try:
        if DEFAULT_DATA_FILE.exists() and DEFAULT_DATA_FILE.stat().st_size > 0:
            return DEFAULT_DATA_FILE.read_bytes(), {"filename": DEFAULT_DATA_FILE.name, "sheet": DEFAULT_SHEET}
    except Exception:
        pass

    return None, {}


def _save_state(state: dict) -> None:
    try:
        APP_STATE_FILE.write_text(json.dumps(state, ensure_ascii=False, indent=2), encoding="utf-8")
    except Exception:
        pass


def _load_state() -> dict:
    try:
        if APP_STATE_FILE.exists():
            return json.loads(APP_STATE_FILE.read_text(encoding="utf-8"))
    except Exception:
        return {}
    return {}

# =========================================================
# Funções de dados
# =========================================================
def br_date_str(dt) -> str:
    if pd.isna(dt):
        return ""
    try:
        return pd.to_datetime(dt).strftime(DATE_FMT_BR)
    except Exception:
        return ""


def normalizar_situacao(x: str) -> str:
    s = str(x).strip().upper()
    if s in ("ATRASADA", "ATRASADO"):
        return "ATRASADA"
    if s in ("NO PRAZO", "NOPRAZO", "NO_PRAZO"):
        return "NO PRAZO"
    if s in ("", "NAN", "NONE"):
        return ""
    return s


@st.cache_data(show_spinner=False)
def carregar_df(upload_bytes: bytes, sheet_name: str = DEFAULT_SHEET) -> pd.DataFrame:
    # Leitura simplificada: o app não apresenta mais o campo "Nome da aba".
    # Tenta abrir a aba padrão; se ela não existir, utiliza automaticamente a primeira aba da planilha.
    try:
        df = pd.read_excel(io.BytesIO(upload_bytes), sheet_name=sheet_name)
    except Exception:
        df = pd.read_excel(io.BytesIO(upload_bytes), sheet_name=0)

    obrigatorias = [COL_CODIGO, COL_TITULO, COL_STATUS, COL_DATA, COL_MOTIVO]
    faltantes = [c for c in obrigatorias if c not in df.columns]
    if faltantes:
        raise ValueError(f"Colunas obrigatórias não encontradas: {', '.join(faltantes)}")

    df = df.copy()
    df[COL_DATA] = pd.to_datetime(df[COL_DATA], errors="coerce", dayfirst=True)
    df = df.dropna(subset=[COL_DATA])

    for col in df.columns:
        if df[col].dtype == "object":
            df[col] = df[col].astype(str).str.strip().replace("nan", "")

    if COL_SITUACAO in df.columns:
        df[COL_SITUACAO] = df[COL_SITUACAO].apply(normalizar_situacao)

    df["Ano"] = df[COL_DATA].dt.year.astype(int)
    df["Mês Nº"] = df[COL_DATA].dt.month.astype(int)
    df["Mês"] = df["Mês Nº"].map(MESES_ABREV)
    df["Mês/Ano"] = df.apply(lambda r: f"{MESES_ABREV.get(int(r['Mês Nº']), r['Mês Nº'])}/{int(r['Ano'])}", axis=1)
    return df


def normalizar_meses_selecionados(meses) -> list[str]:
    """Retorna lista válida de meses abreviados. Se vier vazio, considera todos os meses."""
    todos = [MESES_ABREV[m] for m in range(1, 13)]
    if meses is None:
        return todos
    if isinstance(meses, str):
        meses = [meses]
    meses = [str(m) for m in meses if str(m).strip()]
    # Compatibilidade com versões antigas que tinham a opção (Todos).
    if not meses or "(Todos)" in meses:
        return todos
    validos = [m for m in meses if m in INV_MESES_ABREV]
    return validos or todos


def normalizar_anos_selecionados(anos, anos_disponiveis=None) -> list[int]:
    """Retorna lista de anos válida. Se vier vazio ou (Todos), considera todos os anos disponíveis."""
    if anos_disponiveis is None:
        anos_disponiveis = []
    if anos is None:
        return [int(a) for a in anos_disponiveis]
    if isinstance(anos, (str, int)):
        anos = [anos]
    anos_txt = [str(a).strip() for a in anos if str(a).strip()]
    if not anos_txt or "(Todos)" in anos_txt:
        return [int(a) for a in anos_disponiveis]
    out = []
    for a in anos_txt:
        try:
            out.append(int(a))
        except Exception:
            pass
    return sorted(list(dict.fromkeys(out)))


def filtrar_periodo(df: pd.DataFrame, ano, meses) -> pd.DataFrame:
    dff = df.copy()
    anos_disponiveis = sorted(dff["Ano"].dropna().astype(int).unique().tolist())
    anos_validos = normalizar_anos_selecionados(ano, anos_disponiveis)
    if anos_validos:
        dff = dff[dff["Ano"].isin(anos_validos)]
    meses_validos = normalizar_meses_selecionados(meses)
    meses_num = [INV_MESES_ABREV[m] for m in meses_validos]
    dff = dff[dff["Mês Nº"].isin(meses_num)]
    return dff


def aplicar_filtros_adicionais(df: pd.DataFrame, filtros: dict) -> pd.DataFrame:
    dff = df.copy()
    for col, selecionados in filtros.items():
        if col not in dff.columns:
            continue
        if not selecionados:
            return dff.iloc[0:0]
        dff = dff[dff[col].astype(str).isin(selecionados)]
    return dff


def resumo_periodo(df: pd.DataFrame) -> dict:
    total = int(len(df))
    atrasadas = 0
    if COL_SITUACAO in df.columns and total:
        atrasadas = int((df[COL_SITUACAO].apply(normalizar_situacao) == "ATRASADA").sum())
    p_ini = br_date_str(df[COL_DATA].min()) if total else "-"
    p_fim = br_date_str(df[COL_DATA].max()) if total else "-"
    return {"total": total, "atrasadas": atrasadas, "periodo": f"{p_ini} → {p_fim}"}


def preparar_ocorrencias_normal(df: pd.DataFrame, ano_sel, mes_sel) -> pd.DataFrame:
    """Prepara o gráfico normal para um ou vários anos/meses selecionados."""
    if df.empty:
        return pd.DataFrame({"Período": ["SEM DADOS"], "Ocorrências": [0]})

    anos_validos = normalizar_anos_selecionados(ano_sel, sorted(df["Ano"].dropna().astype(int).unique().tolist()))
    meses_validos = normalizar_meses_selecionados(mes_sel)

    # Quando houver mais de um ano ou mais de um mês, mostrar evolução por mês/ano.
    if len(anos_validos) > 1 or len(meses_validos) > 1:
        base = df.copy()
        g = base.groupby(["Ano", "Mês Nº", "Mês/Ano"], as_index=False).size().rename(columns={"size": "Ocorrências"})
        g["_ordem"] = g["Ano"] * 100 + g["Mês Nº"]
        g = g.sort_values("_ordem")
        return g[["Mês/Ano", "Ocorrências"]].rename(columns={"Mês/Ano": "Período"})

    label = periodo_label(anos_validos, meses_validos)
    return pd.DataFrame({"Período": [label], "Ocorrências": [int(len(df))]})


def preparar_ocorrencias_por_mes_comparacao(df1: pd.DataFrame, label1: str, df2: pd.DataFrame, label2: str, meses_1, meses_2) -> pd.DataFrame:
    """Base auxiliar para exportação: mostra a composição mensal de cada período comparado."""
    linhas = []
    for dfp, label, meses in [(df1, label1, meses_1), (df2, label2, meses_2)]:
        meses_validos = normalizar_meses_selecionados(meses)
        contagem = dfp.groupby("Mês Nº").size() if not dfp.empty else pd.Series(dtype=int)
        for mes in meses_validos:
            n = INV_MESES_ABREV[mes]
            linhas.append({"Período": label, "Mês": mes, "Ocorrências": int(contagem.get(n, 0))})
    return pd.DataFrame(linhas)


def preparar_motivos_normal(df: pd.DataFrame, top_n: int = 12) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame({"Motivo": ["SEM DADOS"], "Ocorrências": [0]})
    g = df[COL_MOTIVO].fillna("").replace("", "SEM MOTIVO").value_counts().head(top_n).reset_index()
    g.columns = ["Motivo", "Ocorrências"]
    return g


def preparar_ocorrencias_comparacao(df1: pd.DataFrame, label1: str, df2: pd.DataFrame, label2: str) -> pd.DataFrame:
    return pd.DataFrame(
        {
            "Período": [label1, label2],
            "Ano": [label1, label2],
            "Ocorrências": [int(len(df1)), int(len(df2))],
        }
    )


def preparar_motivos_comparacao(df1: pd.DataFrame, label1: str, df2: pd.DataFrame, label2: str, top_n: int = 12) -> pd.DataFrame:
    m1 = df1[COL_MOTIVO].fillna("").replace("", "SEM MOTIVO").value_counts()
    m2 = df2[COL_MOTIVO].fillna("").replace("", "SEM MOTIVO").value_counts()
    motivos = (m1.add(m2, fill_value=0).sort_values(ascending=False).head(top_n).index.tolist())
    linhas = []
    for motivo in motivos:
        linhas.append({"Motivo": motivo, "Período": label1, "Ocorrências": int(m1.get(motivo, 0))})
        linhas.append({"Motivo": motivo, "Período": label2, "Ocorrências": int(m2.get(motivo, 0))})
    if not linhas:
        linhas = [
            {"Motivo": "SEM DADOS", "Período": label1, "Ocorrências": 0},
            {"Motivo": "SEM DADOS", "Período": label2, "Ocorrências": 0},
        ]
    return pd.DataFrame(linhas)


def periodo_label(ano, meses) -> str:
    anos_validos = normalizar_anos_selecionados(ano, [])
    if not anos_validos and isinstance(ano, list):
        anos_validos = [int(a) for a in ano if str(a).isdigit()]
    ano_txt = "+".join([str(a) for a in anos_validos]) if anos_validos else str(ano)
    meses_validos = normalizar_meses_selecionados(meses)
    todos = [MESES_ABREV[m] for m in range(1, 13)]
    if meses_validos == todos:
        return f"Ano {ano_txt}"
    if len(meses_validos) == 1:
        return f"{meses_validos[0]}/{ano_txt}"
    return f"{'+'.join(meses_validos)}/{ano_txt}"


def calcular_variacao(valor_base: int, valor_comp: int) -> tuple[int, float]:
    delta = int(valor_comp) - int(valor_base)
    pct = (delta / int(valor_base) * 100) if int(valor_base) else 0.0
    return delta, pct


def _formatar_planilha(ws):
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="1F4E79")
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    # Ajuste compatível com Streamlit Cloud/openpyxl:
    # em planilhas com células mescladas, col[0] pode ser MergedCell e não ter column_letter.
    for col_idx, col in enumerate(ws.iter_cols(), start=1):
        max_len = 10
        letter = get_column_letter(col_idx)
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value)) if cell.value is not None else 0)
            except Exception:
                pass
        ws.column_dimensions[letter].width = min(max_len + 2, 42)


def _adicionar_tabela(ws, df: pd.DataFrame, start_row: int = 1, start_col: int = 1):
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=start_row):
        for c_idx, value in enumerate(row, start=start_col):
            ws.cell(row=r_idx, column=c_idx, value=value)
    _formatar_planilha(ws)


def _formatar_recorte_compacto(ws, largura_padrao: float = 16, altura_padrao: float = 17, fonte_tamanho: int = 9):
    """Formata a aba RECORTE com visual compacto e dimensões fixas.

    Mantém fonte, altura de linha e largura de coluna padronizadas para evitar
    variações causadas por textos longos. Colunas auxiliares são ocultadas.
    """
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="1F4E79")

    for row_idx, row in enumerate(ws.iter_rows(), start=1):
        ws.row_dimensions[row_idx].height = altura_padrao
        for cell in row:
            cell.border = border
            cell.font = Font(size=fonte_tamanho, color="1F2937")
            cell.alignment = Alignment(
                horizontal="left",
                vertical="center",
                wrap_text=False,
                shrink_to_fit=True,
            )

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True, size=fonte_tamanho)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False, shrink_to_fit=True)

    # Dimensão fixa para todas as colunas; evita autoajuste que distorce a tabela.
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = largura_padrao

    # Ajustes mínimos em colunas textuais principais, ainda com tamanho controlado.
    cabecalhos = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
    larguras = {
        COL_CODIGO: 11,
        COL_DATA: 12,
        COL_STATUS: 13,
        COL_SITUACAO: 13,
        COL_TITULO: 24,
        COL_MOTIVO: 24,
        COL_RESP_OCORRENCIA: 20,
        COL_RESP_ANALISE: 22,
        COL_CLIENTE: 20,
        "Embalagem": 15,
        "Período": 18,
    }
    for nome_coluna, largura in larguras.items():
        if nome_coluna in cabecalhos:
            ws.column_dimensions[get_column_letter(cabecalhos[nome_coluna])].width = largura

    # Oculta colunas auxiliares/menos utilizadas no recorte exportado.
    colunas_ocultar = {COL_CATEGORIA, "Local", "ANO", "Ano", "MÊS", "Mês", "MÊS Nº", "Mês Nº", "Mes", "Mes Nº"}
    for nome_coluna, col_idx in cabecalhos.items():
        if nome_coluna in colunas_ocultar:
            ws.column_dimensions[get_column_letter(col_idx)].hidden = True

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _preparar_figura_para_excel(fig, margem_inferior: int = 130):
    """Replica a figura Plotly e aumenta a margem inferior para não cortar rótulos inclinados."""
    try:
        fig_excel = fig.full_copy()
    except Exception:
        try:
            import copy
            fig_excel = copy.deepcopy(fig)
        except Exception:
            fig_excel = fig
    try:
        fig_excel.update_xaxes(tickangle=-45, automargin=True)
        fig_excel.update_layout(margin=dict(l=40, r=30, t=70, b=margem_inferior))
    except Exception:
        pass
    return fig_excel


def _adicionar_grafico_barras(ws, title: str, cat_col: int, val_col: int, start_row: int, end_row: int, anchor: str, series_title_from_header: bool = True):
    if end_row <= start_row:
        return
    chart = BarChart()
    chart.type = "col"
    chart.title = title
    chart.height = 9
    chart.width = 18
    chart.legend = None
    data = Reference(ws, min_col=val_col, min_row=start_row, max_row=end_row)
    cats = Reference(ws, min_col=cat_col, min_row=start_row + 1, max_row=end_row)
    chart.add_data(data, titles_from_data=series_title_from_header)
    chart.set_categories(cats)
    chart.dLbls = DataLabelList()
    chart.dLbls.showVal = True
    chart.dLbls.dLblPos = "outEnd"
    try:
        chart.y_axis.majorGridlines = None
        chart.y_axis.tickLblPos = "none"
    except Exception:
        pass
    try:
        # Inclina os rótulos do eixo X em 45° para evitar corte/sobreposição.
        chart.x_axis.textRotation = 45
        chart.x_axis.tickLblPos = "low"
    except Exception:
        pass
    ws.add_chart(chart, anchor)


def _adicionar_grafico_barras_multiserie(ws, title: str, cat_col: int, first_val_col: int, last_val_col: int, start_row: int, end_row: int, anchor: str):
    """Adiciona gráfico de barras agrupadas com múltiplas séries.

    Usado na comparação por período, mantendo o eixo X como Motivo ou Período
    e cada período/ano como uma série distinta.
    """
    if end_row <= start_row or last_val_col < first_val_col:
        return
    chart = BarChart()
    chart.type = "col"
    chart.title = title
    chart.height = 9
    chart.width = 18
    chart.legend = None if first_val_col == last_val_col else chart.legend
    data = Reference(ws, min_col=first_val_col, max_col=last_val_col, min_row=start_row, max_row=end_row)
    cats = Reference(ws, min_col=cat_col, min_row=start_row + 1, max_row=end_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.dLbls = DataLabelList()
    chart.dLbls.showVal = True
    chart.dLbls.dLblPos = "outEnd"
    try:
        chart.y_axis.majorGridlines = None
        chart.y_axis.tickLblPos = "none"
    except Exception:
        pass
    try:
        chart.x_axis.textRotation = 45
        chart.x_axis.tickLblPos = "low"
    except Exception:
        pass
    ws.add_chart(chart, anchor)


def _escrever_base_grafico(ws, df: pd.DataFrame, start_row: int, start_col: int):
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=start_row):
        for c_idx, value in enumerate(row, start=start_col):
            ws.cell(row=r_idx, column=c_idx, value=value)
    return start_row, start_col, start_row + len(df), start_col + df.shape[1] - 1


def gerar_relatorio_excel(
    df_tabela: pd.DataFrame,
    df_occ: pd.DataFrame,
    df_mot: pd.DataFrame,
    modo_comparacao: bool,
    resumo_df: pd.DataFrame,
    variacao_df: pd.DataFrame | None = None,
    df_occ_mensal: pd.DataFrame | None = None,
    figs_plotly: list | None = None,
) -> bytes:
    wb = Workbook()
    ws_dash = wb.active
    ws_dash.title = "DASHBOARD"
    ws_dash.sheet_view.showGridLines = False

    ws_resumo = wb.create_sheet("RESUMO")
    _adicionar_tabela(ws_resumo, resumo_df)

    ws_occ = wb.create_sheet("DADOS_OCORRENCIAS")
    _adicionar_tabela(ws_occ, df_occ)

    ws_mot = wb.create_sheet("DADOS_MOTIVOS")
    _adicionar_tabela(ws_mot, df_mot)

    if df_occ_mensal is not None and not df_occ_mensal.empty:
        ws_mes = wb.create_sheet("COMPOSICAO_MENSAL")
        _adicionar_tabela(ws_mes, df_occ_mensal)

    if variacao_df is not None:
        ws_var = wb.create_sheet("VARIACAO")
        _adicionar_tabela(ws_var, variacao_df)

    df_rec = df_tabela.copy()
    if COL_DATA in df_rec.columns:
        df_rec[COL_DATA] = df_rec[COL_DATA].apply(br_date_str)
    ws_rec = wb.create_sheet("RECORTE")
    _adicionar_tabela(ws_rec, df_rec)
    _formatar_recorte_compacto(ws_rec)

    # Dashboard sintético com os mesmos dados dos gráficos do app.
    ws_dash["A1"] = "INDICADORES QUALIDADE RS — RELATÓRIO DO RECORTE"
    ws_dash["A1"].font = Font(bold=True, size=16, color="1F4E79")
    ws_dash.merge_cells("A1:H1")
    ws_dash["A2"] = f"VERSÃO {APP_VERSION}"
    ws_dash["A2"].font = Font(italic=True, size=10, color="666666")
    ws_dash.merge_cells("A2:H2")
    ws_dash["A3"] = "Resumo"
    ws_dash["A3"].font = Font(bold=True, size=12)

    # Copia resumo para o dashboard.
    for r_idx, row in enumerate(dataframe_to_rows(resumo_df, index=False, header=True), start=4):
        for c_idx, value in enumerate(row, start=1):
            ws_dash.cell(row=r_idx, column=c_idx, value=value)

    if variacao_df is not None:
        base_row = 4 + len(resumo_df) + 3
        ws_dash.cell(row=base_row, column=1, value="Variação entre períodos").font = Font(bold=True, size=12)
        for r_idx, row in enumerate(dataframe_to_rows(variacao_df, index=False, header=True), start=base_row + 1):
            for c_idx, value in enumerate(row, start=1):
                ws_dash.cell(row=r_idx, column=c_idx, value=value)

    _formatar_planilha(ws_dash)

    # Gráficos em planilhas de dados.
    # DADOS_OCORRENCIAS: Período no eixo X e rótulo sobre cada barra.
    _adicionar_grafico_barras(
        ws_occ,
        "Número de ocorrências",
        df_occ.columns.get_loc("Período") + 1 if "Período" in df_occ.columns else 1,
        df_occ.columns.get_loc("Ocorrências") + 1,
        1,
        len(df_occ) + 1,
        "E2",
    )

    # DADOS_MOTIVOS: Motivo no eixo X. Em comparação, monta base auxiliar
    # agrupada por Período para barras de cores/séries distintas.
    if modo_comparacao and "Período" in df_mot.columns:
        df_mot_chart = (
            df_mot.pivot_table(index="Motivo", columns="Período", values="Ocorrências", aggfunc="sum", fill_value=0)
            .reset_index()
        )
        start_row, start_col = 1, max(8, df_mot.shape[1] + 3)
        r1, c1, r2, c2 = _escrever_base_grafico(ws_mot, df_mot_chart, start_row, start_col)
        _adicionar_grafico_barras_multiserie(
            ws_mot,
            "Motivos das ocorrências",
            cat_col=c1,
            first_val_col=c1 + 1,
            last_val_col=c2,
            start_row=r1,
            end_row=r2,
            anchor=f"{get_column_letter(c2 + 2)}2",
        )
    else:
        _adicionar_grafico_barras(
            ws_mot,
            "Motivos das ocorrências",
            df_mot.columns.get_loc("Motivo") + 1 if "Motivo" in df_mot.columns else 1,
            df_mot.columns.get_loc("Ocorrências") + 1,
            1,
            len(df_mot) + 1,
            "E2",
        )

    # Dashboard: insere os próprios gráficos Plotly apresentados na tela, preservando visual e cores.
    imagens_inseridas = False
    if figs_plotly:
        anchors = ["A14", "J14"]
        for fig, anchor in zip(figs_plotly, anchors):
            try:
                fig_excel = _preparar_figura_para_excel(fig)
                img_bytes = io.BytesIO(fig_excel.to_image(format="png", scale=2))
                img = XLImage(img_bytes)
                img.width = 680
                img.height = 410
                ws_dash.add_image(img, anchor)
                imagens_inseridas = True
            except Exception:
                imagens_inseridas = False
                break

    # Fallback: caso o ambiente não gere PNG pelo kaleido, cria gráficos nativos do Excel.
    aux_row = 35
    ws_dash.cell(row=aux_row, column=1, value="Base gráfico ocorrências").font = Font(bold=True)
    for r_idx, row in enumerate(dataframe_to_rows(df_occ, index=False, header=True), start=aux_row + 1):
        for c_idx, value in enumerate(row, start=1):
            ws_dash.cell(row=r_idx, column=c_idx, value=value)
    occ_val_col = df_occ.columns.get_loc("Ocorrências") + 1
    if not imagens_inseridas:
        _adicionar_grafico_barras(ws_dash, "Número de ocorrências", 1, occ_val_col, aux_row + 1, aux_row + len(df_occ) + 1, "A14")

    aux_mot_row = aux_row + len(df_occ) + 15
    ws_dash.cell(row=aux_mot_row, column=1, value="Base gráfico motivos").font = Font(bold=True)
    for r_idx, row in enumerate(dataframe_to_rows(df_mot, index=False, header=True), start=aux_mot_row + 1):
        for c_idx, value in enumerate(row, start=1):
            ws_dash.cell(row=r_idx, column=c_idx, value=value)
    mot_val_col = df_mot.columns.get_loc("Ocorrências") + 1
    if not imagens_inseridas:
        _adicionar_grafico_barras(ws_dash, "Motivos das ocorrências", 1, mot_val_col, aux_mot_row + 1, aux_mot_row + len(df_mot) + 1, "J14")

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()


# =========================================================
# Manual integrado
# =========================================================
def render_manual_app() -> None:
    st.subheader("📘 Manual de funcionamento do aplicativo")
    st.markdown(
        """
        **Aplicativo de consulta e análise de ocorrências por período e motivo.**

        A base de dados deve ser retirada de consulta no sistema de gestão utilizado pela empresa,
        exportada para Excel e carregada no aplicativo. Após o upload, o sistema mantém a última base
        registrada para consulta automática na próxima abertura.
        """
    )

    st.markdown("### Fluxo decisório de uso")
    st.markdown(
        """
        1. **Preciso analisar ocorrências?**  
           → Exporto a consulta do sistema de gestão para Excel.

        2. **Tenho uma nova base?**  
           → Sim: faço upload da nova planilha.  
           → Não: o app utiliza automaticamente a última base registrada.

        3. **Quero analisar um período único?**  
           → Seleciono ano(s) e mês(es) na consulta normal.

        4. **Quero comparar períodos?**  
           → Ativo a análise comparativa, seleciono o(s) ano(s) do Período 1, os meses de análise e o(s) ano(s) do Período 2.  
           → Os meses do Período 2 são sincronizados automaticamente com os meses do Período 1.

        5. **Preciso investigar uma causa específica?**  
           → Uso o filtro adicional por motivo.

        6. **Preciso formalizar a análise?**  
           → Uso o botão **Gerar relatório**, que exporta tabela, gráficos e variações em Excel.
        """
    )

    st.markdown("### Tela do aplicativo — campos principais")
    st.info(
        "1) Upload da base Excel | 2) Filtros de ano e mês | 3) Comparação entre períodos | "
        "4) Filtro por motivo | 5) Total de ocorrências | 6) Gráfico por período | "
        "7) Gráfico por motivo | 8) Gerar relatório"
    )

    st.markdown("### Leitura dos gráficos")
    st.markdown(
        """
        - **Número de ocorrências:** mostra o volume de ocorrências no período selecionado ou a comparação entre períodos.
        - **Motivos das ocorrências:** mostra os principais motivos registrados na base filtrada.
        - **Variação entre períodos:** em análise comparativa, indica aumento ou redução entre os períodos selecionados.
        """
    )

    if MANUAL_FILE.exists():
        with open(MANUAL_FILE, "rb") as f:
            st.download_button(
                "📥 Baixar manual completo em PDF",
                data=f,
                file_name="Manual_APP_Indicadores_Qualidade_V_RM_2026.pdf",
                mime="application/pdf",
            )
    else:
        st.warning("Manual PDF não encontrado no diretório do app. A orientação resumida permanece disponível nesta aba.")

# =========================================================
# Interface
# =========================================================
st.set_page_config(page_title=APP_NAME, page_icon="📊", layout="wide")
st.markdown(
    f"""
    <style>
    .stApp {{background: linear-gradient(180deg, {COR_FUNDO} 0%, {COR_FUNDO_2} 38%, #FFFFFF 100%); color: {COR_TEXTO};}}
    .block-container {{padding-top: 1rem; padding-bottom: 1rem;}}
    h1, h2, h3, .stMarkdown, label, p {{color: {COR_TEXTO};}}
    div[data-testid="stMetric"] {{background-color: {COR_CARD}; border: 1px solid {COR_BORDA}; border-radius: 16px; padding: 14px; box-shadow: 0 6px 18px rgba(23,61,53,0.08);}}
    div[data-testid="stMetricLabel"] p {{color: {COR_TEXTO}; font-weight: 700;}}
    section[data-testid="stSidebar"] {{background: linear-gradient(180deg, {COR_SIDEBAR} 0%, #FFFFFF 100%); border-right: 1px solid {COR_BORDA};}}
    div[data-testid="stExpander"] {{background-color: rgba(255,255,255,0.72); border-radius: 14px; border: 1px solid {COR_BORDA};}}
    .main-title {{display:flex; align-items:center; gap:18px; margin-bottom: 0.4rem;}}
    .main-title img {{width:130px; border-radius: 8px;}}
    .main-title h1 {{margin:0; color:{COR_TEXTO};}}
    @media (max-width: 768px) {{
        .block-container {{padding-left: 0.75rem; padding-right: 0.75rem;}}
        div[data-testid="stHorizontalBlock"] {{flex-direction: column !important;}}
        .main-title {{flex-direction:column; align-items:flex-start;}}
    }}
    </style>
    """,
    unsafe_allow_html=True,
)

logo_path = LOGO_FILE if LOGO_FILE.exists() else Path("/mnt/data/images.png")
col_logo, col_title = st.columns([0.8, 5])
with col_logo:
    if logo_path.exists():
        st.image(str(logo_path), width=125)
with col_title:
    st.title(APP_NAME)
    st.markdown(
        """
        **APLICATIVO DE CONSULTA E ANÁLISE DE OCORRÊNCIAS POR PERÍODO E MOTIVO.**  
        Em filtros adicionais é possível filtrar as ocorrências por motivo do período selecionado.
        """
    )
    st.markdown(f"<div style='font-size:10px; font-style:italic; font-weight:600; letter-spacing:0.04em;'>VERSÃO {APP_VERSION}</div>", unsafe_allow_html=True)

with st.expander("📘 Manual", expanded=False):
    render_manual_app()

saved_state = _load_state()
last_bytes, last_meta = _load_last_upload()

with st.sidebar:
    st.header("📥 Base de dados")
    sheet = DEFAULT_SHEET
    up = st.file_uploader("Enviar nova base Excel", type=["xlsx", "xlsm", "xls"])

    if last_bytes and up is None:
        st.success(f"Base carregada automaticamente: {last_meta.get('filename', 'ultimo.xlsx')}")
    elif up is not None:
        st.success(f"Novo upload recebido: {up.name}")
    else:
        st.info("Envie uma base Excel para iniciar.")

if up is not None:
    upload_bytes = up.getvalue()
    upload_name = up.name
else:
    upload_bytes = last_bytes
    upload_name = last_meta.get("filename", "ultimo.xlsx") if last_bytes else None

if not upload_bytes or len(upload_bytes) == 0:
    st.warning("Nenhuma base disponível. Faça upload da consulta exportada do sistema de gestão.")
    st.stop()

try:
    df_base = carregar_df(upload_bytes, sheet)
    _save_last_upload(upload_bytes, upload_name or "ultimo.xlsx", sheet)
except Exception as e:
    st.error(f"Erro ao carregar a base: {e}")
    st.stop()

anos = sorted([int(a) for a in df_base["Ano"].dropna().unique().tolist() if int(a) >= 2025])
if not anos:
    st.warning("A base não possui registros a partir de 2025.")
    st.stop()

mes_options = [MESES_ABREV[m] for m in range(1, 13)]
ano_options = [str(a) for a in anos]

st.subheader("Filtros")
st.caption("Ao abrir o aplicativo, a consulta inicia sem comparação, considerando todos os anos e meses existentes na última base carregada.")
modo_comparacao = st.toggle(
    "Ativar análise comparativa entre períodos",
    value=False,
    key="modo_comparacao_toggle",
)

if not modo_comparacao:
    c1, c2 = st.columns(2)
    with c1:
        ano_sel = st.multiselect(
            "Ano(s)",
            options=ano_options,
            default=ano_options,
            key="ano_sel_normal",
            help="Por padrão, todos os anos da base ficam selecionados.",
        )
    with c2:
        mes_sel = st.multiselect(
            "Mês(es)",
            options=mes_options,
            default=mes_options,
            key="mes_sel_normal",
            help="Por padrão, todos os meses ficam selecionados. Remova os meses que não deseja analisar.",
        )
else:
    st.markdown("**Período 1 x Período 2**")
    st.caption("Selecione os meses somente no Período 1. O Período 2 usa automaticamente os mesmos meses, mudando apenas o(s) ano(s) de referência.")
    meses_multiselect = mes_options
    c1, c2, c3, c4 = st.columns([1, 2, 1, 2])
    with c1:
        ano_1 = st.multiselect(
            "Ano(s) — Período 1",
            ano_options,
            default=[ano_options[0]],
            key="ano_1_comp",
        )
    with c2:
        mes_1 = st.multiselect(
            "Meses de comparação",
            meses_multiselect,
            default=["Jan"],
            key="mes_1_comp",
        )
    with c3:
        ano_2 = st.multiselect(
            "Ano(s) — Período 2",
            ano_options,
            default=[ano_options[-1]],
            key="ano_2_comp",
        )
    with c4:
        mes_2 = list(mes_1) if mes_1 else meses_multiselect
        st.write("**Meses — Período 2 (automático)**")
        st.info(", ".join(mes_2))
        st.caption("Meses sincronizados com o Período 1.")

with st.expander("Filtros adicionais", expanded=False):
    filtros = {}
    # Categoria removida dos filtros adicionais. Motivo incluído para permitir consulta direta por motivo no período selecionado.
    cols_filtro = [c for c in [COL_STATUS, COL_MOTIVO, COL_CLIENTE, COL_RESP_OCORRENCIA, COL_RESP_ANALISE, COL_SITUACAO] if c in df_base.columns]
    grid = st.columns(3)
    for i, col in enumerate(cols_filtro):
        valores = sorted([v for v in df_base[col].dropna().astype(str).replace("nan", "").unique().tolist() if v != ""])
        default = valores
        with grid[i % 3]:
            filtros[col] = st.multiselect(col, options=valores, default=default)

# Aplica filtros adicionais antes do recorte por período, para todos os gráficos respeitarem o mesmo escopo.
df_escopo = aplicar_filtros_adicionais(df_base, filtros)

if not modo_comparacao:
    df_periodo = filtrar_periodo(df_escopo, ano_sel, mes_sel)
    resumo = resumo_periodo(df_periodo)

    k1, k2, k3, k4 = st.columns(4)
    k1.metric("Ocorrências", resumo["total"])
    k2.metric("Em atraso", resumo["atrasadas"])
    k3.metric("Período", resumo["periodo"])
    k4.metric("Base ativa", upload_name or "último upload")

    st.divider()
    st.subheader("Número de ocorrências")
    df_occ = preparar_ocorrencias_normal(df_periodo, ano_sel, mes_sel)
    fig1 = px.bar(df_occ, x="Período", y="Ocorrências", text="Ocorrências", title=f"Ocorrências — {periodo_label(ano_sel, mes_sel)}")
    fig1.update_traces(marker_color=COR_NORMAL, textposition="outside", cliponaxis=False)
    fig1.update_yaxes(title=None, showticklabels=False)
    fig1.update_layout(showlegend=False, height=430, margin=dict(l=10, r=10, t=55, b=10))
    st.plotly_chart(fig1, use_container_width=True)

    st.subheader("Motivos das ocorrências")
    df_mot = preparar_motivos_normal(df_periodo)
    fig2 = px.bar(df_mot, x="Motivo", y="Ocorrências", text="Ocorrências", title="Motivos — Top 12")
    fig2.update_traces(marker_color=COR_NORMAL, textposition="outside", cliponaxis=False)
    fig2.update_layout(xaxis_tickangle=-45, showlegend=False, height=460, margin=dict(l=10, r=10, t=55, b=10))
    fig2.update_yaxes(title=None, showticklabels=False)
    st.plotly_chart(fig2, use_container_width=True)

    df_tabela = df_periodo.sort_values(COL_DATA, ascending=False)
    resumo_export = pd.DataFrame([{"Período": periodo_label(ano_sel, mes_sel), "Ocorrências": resumo["total"], "Atrasadas": resumo["atrasadas"], "Intervalo": resumo["periodo"]}])
    df_variacao = None
    df_occ_mensal = None
    state_to_save = {"sheet": sheet, "modo_comparacao": False, "ano_sel": ano_sel, "mes_sel": mes_sel, "filtros": filtros}

else:
    df_p1 = filtrar_periodo(df_escopo, ano_1, mes_1)
    df_p2 = filtrar_periodo(df_escopo, ano_2, mes_2)
    label1 = periodo_label(ano_1, mes_1)
    label2 = periodo_label(ano_2, mes_2)

    r1 = resumo_periodo(df_p1)
    r2 = resumo_periodo(df_p2)
    variacao, variacao_pct = calcular_variacao(r1["total"], r2["total"])
    variacao_atrasadas, variacao_atrasadas_pct = calcular_variacao(r1["atrasadas"], r2["atrasadas"])
    df_variacao = pd.DataFrame([
        {"Indicador": "Ocorrências", "Período 1": r1["total"], "Período 2": r2["total"], "Variação absoluta": variacao, "Variação %": round(variacao_pct, 1)},
        {"Indicador": "Atrasadas", "Período 1": r1["atrasadas"], "Período 2": r2["atrasadas"], "Variação absoluta": variacao_atrasadas, "Variação %": round(variacao_atrasadas_pct, 1)},
    ])
    df_occ_mensal = preparar_ocorrencias_por_mes_comparacao(df_p1, label1, df_p2, label2, mes_1, mes_2)

    k1, k2, k3, k4 = st.columns(4)
    k1.metric(label1, r1["total"])
    k2.metric(label2, r2["total"], delta=variacao)
    k3.metric("Variação %", f"{variacao_pct:.1f}%")
    k4.metric("Base ativa", upload_name or "último upload")

    with st.expander("Dados da variação entre períodos", expanded=False):
        st.dataframe(df_variacao, use_container_width=True, hide_index=True)

    st.divider()
    st.subheader("Número de ocorrências")
    df_occ = preparar_ocorrencias_comparacao(df_p1, label1, df_p2, label2)
    fig1 = px.bar(
        df_occ,
        x="Período",
        y="Ocorrências",
        color="Período",
        text="Ocorrências",
        title=f"Comparação de ocorrências — {label1} x {label2}",
        color_discrete_sequence=CORES_COMPARACAO,
    )
    fig1.update_traces(textposition="outside", cliponaxis=False)
    fig1.update_yaxes(title=None, showticklabels=False)
    fig1.update_layout(showlegend=True, height=430, margin=dict(l=10, r=10, t=55, b=10))
    st.plotly_chart(fig1, use_container_width=True)

    st.subheader("Motivos das ocorrências")
    df_mot = preparar_motivos_comparacao(df_p1, label1, df_p2, label2)
    fig2 = px.bar(
        df_mot,
        x="Motivo",
        y="Ocorrências",
        color="Período",
        barmode="group",
        text="Ocorrências",
        title=f"Comparação dos motivos — {label1} x {label2}",
        color_discrete_sequence=CORES_COMPARACAO,
    )
    fig2.update_traces(textposition="outside", cliponaxis=False)
    fig2.update_layout(xaxis_tickangle=-45, showlegend=True, height=500, margin=dict(l=10, r=10, t=55, b=10))
    fig2.update_yaxes(title=None, showticklabels=False)
    st.plotly_chart(fig2, use_container_width=True)

    df_tabela = pd.concat([
        df_p1.assign(Período=label1),
        df_p2.assign(Período=label2),
    ], ignore_index=True).sort_values(COL_DATA, ascending=False)
    resumo_export = pd.DataFrame([
        {"Período": label1, "Ocorrências": r1["total"], "Atrasadas": r1["atrasadas"], "Intervalo": r1["periodo"]},
        {"Período": label2, "Ocorrências": r2["total"], "Atrasadas": r2["atrasadas"], "Intervalo": r2["periodo"]},
    ])
    state_to_save = {
        "sheet": sheet,
        "modo_comparacao": True,
        "ano_1": ano_1,
        "mes_1": mes_1,
        "ano_2": ano_2,
        "mes_2": mes_2,
        "filtros": filtros,
    }

_save_state(state_to_save)

st.divider()
with st.expander("Tabela do recorte", expanded=False):
    st.dataframe(df_tabela, use_container_width=True, height=360)
    st.download_button(
        "📥 Gerar relatório",
        data=gerar_relatorio_excel(
            df_tabela=df_tabela,
            df_occ=df_occ,
            df_mot=df_mot,
            modo_comparacao=modo_comparacao,
            resumo_df=resumo_export,
            variacao_df=df_variacao if modo_comparacao else None,
            df_occ_mensal=df_occ_mensal if modo_comparacao else None,
            figs_plotly=[fig1, fig2],
        ),
        file_name=f"relatorio_indicadores_qualidade_{APP_VERSION}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
